import logging
import multiprocessing
import os
import random
import time
from abc import ABC, abstractmethod
from datetime import datetime, timedelta
from functools import lru_cache
from pathlib import Path
from typing import Dict, Iterator, List, Optional

import numpy as np
import yfinance as yf
from dask.distributed import Client, LocalCluster
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field, PrivateAttr
from tqdm import tqdm
from yfinance.exceptions import YFException, YFRateLimitError

# --- SETUP ---
os.environ["MODIN_ENGINE"] = "dask"
os.environ["DASK_DISTRIBUTED__CLIENT__HEARTBEAT"] = "90s"
os.environ["DASK_DISTRIBUTED__WORKER__MEMORY_LIMIT"] = "0.8"

import modin.pandas as pd  # Import Modin after setting environment variables

os.environ.update({
    k: str(multiprocessing.cpu_count())
    for k in ["MKL_NUM_THREADS", "NUMEXPR_NUM_THREADS", "OMP_NUM_THREADS"]
})

# Constants
NOT_AVAILABLE = "N/A"


class Config(BaseModel):
    """Configuration class to manage constants and settings."""

    DAYS_IN_YEAR: float = 365.25
    INITIAL_CASH: float = 10_000.0
    COMMISSION: float = 0.002
    FONT_SIZE: int = 16
    EPS_THRESHOLD: float = 2.0
    CAGR_THRESHOLD: float = 0.125
    SHARPE_RATIO_THRESHOLD: float = 0.875
    MAX_PE_RATIO: float = 50.0
    MIN_END_PRICE: float = 7.50
    MAX_END_PRICE: float = 250.00
    MAX_EV_EBITDA: float = 37.5
    MIN_ROE: float = 0.1250
    BATCH_SIZE: int = Field(
        default=max(1,
                    multiprocessing.cpu_count() // 2),
        description="Batch size for processing",
    )
    LOOKBACK_PERIOD_YEARS: int = 2
    EXPORT_FILE_PATH: Path = Path("qualified_stocks.xlsx")
    SYMBOLS_FILE_PATH: Path = Path("all_tickers.txt")
    RETRY_ATTEMPTS: int = 2
    RETRY_DELAY_MIN: float = 0.5
    RATE_LIMIT_DELAY: float = 0.5
    MAX_BACKOFF_TIME: float = 30.0
    LOG_LEVEL: int = logging.INFO
    PROCESS_BATCH_SIZE: int = 5

    class Config:
        arbitrary_types_allowed = True


def configure_logger(level: int) -> logging.Logger:
    """Configures and returns a logger for the application.

    Args:
        level: The logging level (e.g., logging.INFO, logging.DEBUG).

    Returns:
        A configured logger instance.
    """
    logger = logging.getLogger(__name__)
    logger.setLevel(level)
    handler = logging.StreamHandler()
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    logging.getLogger("yfinance").setLevel(
        logging.ERROR)  # Suppress yfinance logs
    return logger


logger = configure_logger(Config().LOG_LEVEL)


def log_error(logger, message: str, exception: Exception):
    """Logs error messages in a consistent format.

    Args:
        logger: The logger instance to use.
        message: The base error message.
        exception: The exception that occurred.
    """
    logger.error(f"{message}: {str(exception)}")


class DataFetcher(ABC):
    """Abstract base class for data fetchers."""

    @abstractmethod
    def fetch_data(self, symbol: str) -> Optional[Dict]:
        """Fetches general stock data for a given symbol.

        Args:
            symbol: The stock symbol.

        Returns:
            A dictionary containing stock info, or None.
        """
        pass

    @abstractmethod
    def fetch_stock_data(self, symbol: str, start_date: datetime,
                         end_date: datetime) -> Optional[pd.DataFrame]:
        """Fetches historical stock data for a given symbol.

        Args:
            symbol: The stock symbol.
            start_date: start date.
            end_date: end date.

        Returns:
             pandas Dataframe or None
        """
        pass


class YahooFinanceFetcher(DataFetcher):
    """Fetcher class to interact with Yahoo Finance API."""

    def __init__(self):
        self.config = Config()

    @lru_cache(maxsize=1024)  # Cache up to 1024 symbols
    def fetch_data(self, symbol: str) -> Optional[Dict]:
        """Fetches data for a given symbol from Yahoo Finance.

        Args:
            symbol: The stock symbol to fetch data for.

        Returns:
            stock info or none

        Raises:
            YFRateLimitError: If the Yahoo Finance rate limit exceeded.
            YFException: If yfinance has issue.
            Exception: general errors
        """
        return self.fetch_with_retry(self._fetch_data_inner, symbol)

    def _fetch_data_inner(self, symbol: str) -> Optional[Dict]:
        ticker = yf.Ticker(symbol)
        data = ticker.info
        if not data:
            logger.debug(f"No data available for {symbol}")
            return None
        return data

    def fetch_stock_data(self, symbol: str, start_date: datetime,
                         end_date: datetime) -> Optional[pd.DataFrame]:
        """Fetches historical stock data.

        Args:
            symbol: symbol.
            start_date: date.
            end_date: date.

        Returns:
            Stock data Dataframe
        Raises:
            YFRateLimitError: If the Yahoo Finance rate limit is exceeded.
            YFException: If yfinance encounters an unexpected error.
            Exception: For any other unexpected errors.
        """
        return self.fetch_with_retry(self._fetch_stock_data_inner, symbol,
                                     start_date, end_date)

    def _fetch_stock_data_inner(self, symbol: str, start_date: datetime,
                                end_date: datetime) -> Optional[pd.DataFrame]:
        stock_data = yf.download(
            symbol,
            start=start_date,
            end=end_date,
            progress=False,
            auto_adjust=True,
        )
        if stock_data.empty:
            logger.debug(f"Insufficient stock data for {symbol}")
            return None

        if isinstance(stock_data.columns, pd.MultiIndex):
            stock_data.columns = stock_data.columns.droplevel(1)

        return stock_data

    def fetch_with_retry(self, fetch_func, *args, **kwargs):
        """Fetches data with retries.

        Args:
            fetch_func: function.
            *args: Positional arguments to pass to fetch_func.
            **kwargs: Keyword arguments to pass to fetch_func.

        Returns:
            result of fetch_func

        Raises:
            YFRateLimitError: rate limit
            YFException: yfinance errors.
            Exception: other errors.
        """
        attempt = 0
        config = Config()
        while attempt < config.RETRY_ATTEMPTS:
            try:
                return fetch_func(*args, **kwargs)
            except YFRateLimitError as e:
                log_error(logger, "Rate limit hit, retrying", e)
                wait_time = min(
                    config.MAX_BACKOFF_TIME,
                    config.RETRY_DELAY_MIN * (2**attempt) +
                    random.uniform(0, 1),
                )
                time.sleep(wait_time)
                attempt += 1
                logger.info(f"Retry attempt {attempt} for {args[0]}")
            except YFException as e:
                if "No data found" in str(e):
                    logger.debug("No data found")
                    return None
                log_error(logger, "YFException, retrying", e)
                wait_time = min(
                    config.MAX_BACKOFF_TIME,
                    config.RETRY_DELAY_MIN * (2**attempt) +
                    random.uniform(0, 1),
                )
                time.sleep(wait_time)
                attempt += 1
                logger.info(f"Retry attempt {attempt} for {args[0]}")
            except Exception as e:
                log_error(logger, "Unexpected error", e)
                return None
        return None


class MetricsCalculator(ABC):
    """Abstract Base Class for Metrics Calculators"""

    @abstractmethod
    def calculate_cagr(self, stock_data: pd.DataFrame, start_date: datetime,
                       end_date: datetime) -> float:
        """calculate CAGR

        Args:
            stock_data: dataframe, Close prices.
            start_date: start date.
            end_date: end date.
         Returns:
            CAGR
        """
        pass

    @abstractmethod
    def calculate_sharpe_ratio(self, stock_data: pd.DataFrame) -> float:
        """Calculate sharpe ratio

        Args:
           stock_data: dataframe with Close
        Returns
           sharpe ratio.
        """
        pass


class StockMetricsCalculator(MetricsCalculator):
    """Calculates various stock metrics."""

    @staticmethod
    def calculate_cagr(stock_data: pd.DataFrame, start_date: datetime,
                       end_date: datetime) -> float:
        """Calculates CAGR.

        Args:
            stock_data: Dataframe with Close.
            start_date: start date.
            end_date: end date.

        Returns:
            CAGR.
        """
        start_price = stock_data["Close"].iloc[0]
        end_price = stock_data["Close"].iloc[-1]
        years = (end_date - start_date).days / Config().DAYS_IN_YEAR
        return round((end_price / start_price)**(1 / years) - 1, 3)

    @staticmethod
    def calculate_sharpe_ratio(stock_data: pd.DataFrame) -> float:
        """Calculates Sharpe Ratio.

        Args:
          stock_data: Dataframe with close

        Returns:
           Sharpe ratio.
        """
        returns = stock_data["Close"].pct_change().dropna()
        return (round(returns.mean() / returns.std() *
                      np.sqrt(252), 3) if returns.std() != 0 else 0.0)


def safe_float(value, default=None):
    """Safely converts value to float.

    Args:
        value: The value to convert.
        default: return if fail

    Returns:
      float
    """
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def calculate_ev_ebitda(stock_info: Dict) -> Optional[float]:
    """Calculates the Enterprise Value to EBITDA ratio.

    Args:
        stock_info: Dictionary with 'enterpriseValue' and 'ebitda'.

    Returns:
      EV/EBITDA
    """
    enterprise_value = stock_info.get("enterpriseValue")
    ebitda = stock_info.get("ebitda")
    if enterprise_value and ebitda and ebitda > 0:
        return round(enterprise_value / ebitda, 3)
    return None


class ResultExporter(BaseModel):
    """Handles exporting results to Excel."""

    file_path: Path

    def export_to_excel(self, results: List[Dict[str, float]]) -> None:
        """Exports results

        Args:
            results: result set

        Raises:
           Exception: writing fails
        """
        if not results:
            logger.warning("No results to export; skipping file creation.")
            return

        valid_results = [r for r in results if r is not None]
        if not valid_results:
            logger.warning("No valid results to export")
            return

        required_columns = [
            "symbol",
            "full_name",
            "sector",
            "industry",
            "eps",
            "cagr",
            "roe",
            "pe_ratio",
            "ev_ebitda",
            "sharpe_ratio",
            "end_price",
        ]

        df = pd.DataFrame(valid_results).round(3)
        for col in required_columns:
            if col not in df.columns:
                df[col] = NOT_AVAILABLE
        df = df[required_columns]

        try:
            df.to_excel(self.file_path, index=False)
            self._apply_formatting()
            logger.info(f"Results exported to {self.file_path}")
        except Exception as e:
            log_error(logger, "Failed to export results", e)

    def _apply_formatting(self) -> None:
        """Applies formatting."""
        wb = load_workbook(self.file_path)
        ws = wb.active

        # --- Styling ---
        header_font = Font(size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        data_font_positive = Font(size=12, color="008000")
        data_font_negative = Font(size=12, color="FF0000")
        data_font_neutral = Font(size=12, color="000000")
        data_row_fill = PatternFill("solid", fgColor="D3D3D3")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Header Row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.value = str(cell.value).upper()

        # Data Rows
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = data_row_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

                if cell.column in [5, 6, 7, 8, 9, 10]:  # Columns E-J
                    try:
                        value = float(cell.value)
                        if value > 0:
                            cell.font = data_font_positive
                        elif value < 0:
                            cell.font = data_font_negative
                        else:
                            cell.font = data_font_neutral
                    except (ValueError, TypeError):
                        cell.font = data_font_neutral
                else:
                    cell.font = data_font_neutral

        # Column Widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except ValueError:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(
                12, adjusted_width)

        wb.save(self.file_path)


class StockProcessor(BaseModel):
    """Processes stock symbols, calculates metrics, exports."""

    start_date: datetime
    end_date: datetime
    config: Config = Field(default_factory=Config)
    exporter: ResultExporter = Field(default_factory=lambda: ResultExporter(
        file_path=Config().EXPORT_FILE_PATH))
    fetcher: DataFetcher = Field(default_factory=YahooFinanceFetcher)
    calculator: MetricsCalculator = Field(
        default_factory=StockMetricsCalculator)
    _client: Client = PrivateAttr()  # Dask client

    class Config:
        arbitrary_types_allowed = True

    def process_symbols(self, symbols: List[str]) -> List[Dict[str, float]]:
        """Processes list of symbols.

        Args:
            symbols: list of symbols.

        Returns:
            list of dictionaries.
        """
        results: List[Dict[str, float]] = []
        batch_size = self.config.PROCESS_BATCH_SIZE
        config_dict = self.config.model_dump()

        with tqdm(total=len(symbols), desc="Processing Symbols") as pbar:
            for i in range(0, len(symbols), batch_size):
                batch = symbols[i:i + batch_size]
                futures = self._client.map(
                    StockProcessor.process_symbol_static,
                    batch,
                    [self.start_date] * len(batch),
                    [self.end_date] * len(batch),
                    [config_dict] * len(batch),
                )

                for future in futures:
                    try:
                        result = future.result()
                        if result:
                            results.append(result)
                    except Exception as e:
                        logger.error(f"Error retrieving result: {e}")
                pbar.update(len(batch))
        return results

    @staticmethod
    def process_symbol_static(symbol: str, start_date: datetime,
                              end_date: datetime,
                              config: Dict) -> Optional[Dict[str, float]]:
        """Processes one symbol.

        Args:
            symbol: symbol.
            start_date: date.
            end_date: date.
            config: config dict

        Returns:
            dictionary or None.
        """
        fetcher = YahooFinanceFetcher()
        calculator = StockMetricsCalculator()

        stock_data = fetcher.fetch_stock_data(symbol, start_date, end_date)
        if stock_data is None or stock_data.empty:
            logger.debug(f"No stock data available for {symbol}")
            return None

        stock_info = fetcher.fetch_data(symbol)
        if stock_info is None:
            logger.debug(f"No info available for {symbol}")
            return None

        cagr = calculator.calculate_cagr(stock_data, start_date, end_date)
        sharpe_ratio = calculator.calculate_sharpe_ratio(stock_data)
        eps = safe_float(stock_info.get("trailingEps"))
        pe_ratio = safe_float(stock_info.get("trailingPE"))
        end_price = stock_data["Close"].iloc[-1]
        roe = safe_float(stock_info.get("returnOnEquity"))
        ev_ebitda = calculate_ev_ebitda(stock_info)

        if not all([
                cagr >= config["CAGR_THRESHOLD"],
                eps >= config["EPS_THRESHOLD"],
                sharpe_ratio >= config["SHARPE_RATIO_THRESHOLD"],
                config["MIN_END_PRICE"] <= end_price <= config["MAX_END_PRICE"],
                pe_ratio <= config["MAX_PE_RATIO"],
                roe >= config["MIN_ROE"],
            (ev_ebitda is None or ev_ebitda <= config["MAX_EV_EBITDA"]),
        ]):
            return None

        return {
            "symbol": symbol,
            "full_name": stock_info.get("longName", NOT_AVAILABLE),
            "sector": stock_info.get("sector", NOT_AVAILABLE),
            "industry": stock_info.get("industry", NOT_AVAILABLE),
            "eps": eps,
            "cagr": cagr,
            "sharpe_ratio": sharpe_ratio,
            "pe_ratio": pe_ratio,
            "end_price": round(end_price, 3),
            "ev_ebitda": ev_ebitda,
            "roe": roe,
        }

    def run(self) -> None:
        """Main runner."""
        symbols = self.load_symbols(self.config.SYMBOLS_FILE_PATH)
        if not symbols:
            logger.error("No symbols to process. Exiting.")
            return

        try:
            with Client(
                    LocalCluster(
                        n_workers=self.config.BATCH_SIZE,
                        threads_per_worker=1,
                        processes=True,
                        silence_logs=logging.ERROR,
                    )) as self._client:

                results = self.process_symbols(list(symbols))
                if results:
                    self.exporter.export_to_excel(results)
                else:
                    logger.info("No qualified stocks found.")
        except Exception as e:
            logger.error(f"An error occurred during processing: {e}")

    @staticmethod
    def load_symbols(file_path: Path) -> Iterator[str]:
        """Loads stock symbols from file.

        Args:
            file_path: path

        Returns:
          iterator
        Raises:
            FileNotFoundError: file not found
            Exception: general error
        """
        try:
            with file_path.open("r") as file:
                for line in file:
                    stripped_line = line.strip()
                    if stripped_line:
                        yield stripped_line
        except FileNotFoundError:
            logger.error(f"Symbols file not found: {file_path}")
            return  # Correctly return an empty iterator, not None
        except Exception as e:
            logger.error(f"Error loading symbols {file_path}: {e}")
            return


if __name__ == "__main__":
    config = Config()
    end_date_ = datetime(2024, 12, 31)
    start_date_ = end_date_ - timedelta(days=config.LOOKBACK_PERIOD_YEARS *
                                        config.DAYS_IN_YEAR)

    processor = StockProcessor(start_date=start_date_,
                               end_date=end_date_,
                               config=config)
    processor.run()
